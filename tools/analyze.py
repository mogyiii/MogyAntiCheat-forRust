"""
MogyAntiCheat Player Analysis Tool
====================================
Reads the CSV exported by /ac-export csv and generates:
  - Scatter plots showing "islands" separating normal vs suspicious players
  - Excel workbook with multiple chart sheets
  - PNG images for each chart

Install dependencies:
    pip install pandas matplotlib seaborn openpyxl xlsxwriter scikit-learn

Usage:
    python analyze.py <path_to_csv>
    python analyze.py MogyAntiCheat_Export_20260517_123456.csv
    python analyze.py MogyAntiCheat_Export_20260517_123456.csv --output my_report
"""

import sys
import os
import argparse
import math
from pathlib import Path

try:
    import pandas as pd
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches
    import seaborn as sns
    import numpy as np
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import ScatterChart, Reference, Series
    from openpyxl.chart.series import DataPoint
except ImportError as e:
    print(f"Missing dependency: {e}")
    print("Run: pip install pandas matplotlib seaborn openpyxl xlsxwriter scikit-learn")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

NERF_COLORS = {
    "clean":      ("#4caf50", "Normál (nincs nerf)"),
    "light":      ("#ffeb3b", "Enyhe (1–30%)"),
    "moderate":   ("#ff9800", "Mérsékelt (30–60%)"),
    "heavy":      ("#f44336", "Súlyos (>60%)"),
    "override":   ("#9c27b0", "Kézi felülbírálat"),
}

plt.rcParams.update({
    "figure.facecolor": "#1e1e2e",
    "axes.facecolor":   "#2a2a3e",
    "axes.edgecolor":   "#555577",
    "text.color":       "#e0e0f0",
    "axes.labelcolor":  "#e0e0f0",
    "xtick.color":      "#aaaacc",
    "ytick.color":      "#aaaacc",
    "grid.color":       "#3a3a5a",
    "grid.linestyle":   "--",
    "grid.alpha":       0.5,
})


# ---------------------------------------------------------------------------
# Data loading & preprocessing
# ---------------------------------------------------------------------------

def load_csv(path: str) -> pd.DataFrame:
    df = pd.read_csv(path)
    df.columns = [c.strip() for c in df.columns]

    # Ensure numeric columns
    num_cols = [
        "accuracy", "shots", "hits", "global_nerf",
        "kills", "deaths", "assists",
        "ping_avg", "ping_stddev", "ping_anomalies", "ls_incidents",
    ]
    for col in num_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # manual_override: empty = no override, otherwise it's a % number
    if "manual_override" in df.columns:
        df["has_override"] = df["manual_override"].astype(str).str.strip().ne("") & \
                             df["manual_override"].astype(str).str.strip().ne("nan")
        df["override_pct"] = pd.to_numeric(df["manual_override"], errors="coerce").fillna(0)
    else:
        df["has_override"] = False
        df["override_pct"] = 0.0

    # Derived columns
    df["nerf_pct"] = (1.0 - df["global_nerf"]) * 100.0
    df["kdr"] = df.apply(lambda r: r["kills"] / r["deaths"] if r["deaths"] > 0 else r["kills"], axis=1)
    df["accuracy_pct"] = df["accuracy"] * 100.0
    df["suspicion_level"] = df["nerf_pct"].apply(classify_suspicion)

    return df


def classify_suspicion(nerf_pct: float) -> str:
    if nerf_pct <= 0:
        return "clean"
    elif nerf_pct <= 30:
        return "light"
    elif nerf_pct <= 60:
        return "moderate"
    else:
        return "heavy"


def player_summary(df: pd.DataFrame) -> pd.DataFrame:
    """Aggregate per-player stats across all weapons."""
    agg = df.groupby("player_id").agg(
        avg_accuracy=("accuracy_pct", "mean"),
        max_accuracy=("accuracy_pct", "max"),
        total_shots=("shots", "sum"),
        total_hits=("hits", "sum"),
        kills=("kills", "first"),
        deaths=("deaths", "first"),
        assists=("assists", "first"),
        ping_avg=("ping_avg", "first"),
        ping_stddev=("ping_stddev", "first"),
        ping_anomalies=("ping_anomalies", "first"),
        ls_incidents=("ls_incidents", "first"),
        global_nerf=("global_nerf", "min"),
        has_override=("has_override", "first"),
    ).reset_index()

    agg["nerf_pct"] = (1.0 - agg["global_nerf"]) * 100.0
    agg["kdr"] = agg.apply(lambda r: r["kills"] / r["deaths"] if r["deaths"] > 0 else r["kills"], axis=1)
    agg["overall_accuracy"] = agg.apply(
        lambda r: (r["total_hits"] / r["total_shots"] * 100) if r["total_shots"] > 0 else 0, axis=1
    )
    agg["suspicion_level"] = agg["nerf_pct"].apply(classify_suspicion)
    return agg


# ---------------------------------------------------------------------------
# Chart helpers
# ---------------------------------------------------------------------------

def suspicion_color(level: str) -> str:
    return NERF_COLORS.get(level, NERF_COLORS["clean"])[0]


def dot_size(shots: float, min_size: int = 40, max_size: int = 400) -> float:
    return min(max_size, max(min_size, math.log(shots + 1, 2) * 20))


def add_legend(ax, levels_present: list):
    handles = []
    for lvl in levels_present:
        color, label = NERF_COLORS.get(lvl, ("#ffffff", lvl))
        handles.append(mpatches.Patch(color=color, label=label))
    ax.legend(handles=handles, loc="upper left", framealpha=0.3,
              facecolor="#2a2a3e", edgecolor="#555577", fontsize=9)


def add_annotations(ax, df_player: pd.DataFrame, x_col: str, y_col: str, threshold: float = 0.85):
    """Label outlier points (top suspicion + high values)."""
    heavy = df_player[df_player["suspicion_level"] == "heavy"]
    if heavy.empty:
        return
    p90_x = df_player[x_col].quantile(threshold)
    p90_y = df_player[y_col].quantile(threshold)
    notable = heavy[(heavy[x_col] >= p90_x) | (heavy[y_col] >= p90_y)]
    for _, row in notable.iterrows():
        ax.annotate(
            str(row["player_id"])[-6:],
            (row[x_col], row[y_col]),
            textcoords="offset points", xytext=(6, 6),
            fontsize=7, color="#ff9090", alpha=0.8,
        )


def save_fig(fig, out_dir: Path, name: str) -> Path:
    path = out_dir / f"{name}.png"
    fig.savefig(path, dpi=150, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)
    return path


# ---------------------------------------------------------------------------
# Individual charts
# ---------------------------------------------------------------------------

def chart_accuracy_vs_weighted(df_player: pd.DataFrame, out_dir: Path) -> Path:
    """Main 'island' chart: overall accuracy vs max single-weapon accuracy."""
    fig, ax = plt.subplots(figsize=(10, 7))
    fig.suptitle("Pontosság: Átlag vs Maximum (fegyverenként)", color="#e0e0f0", fontsize=13)

    for level in ["clean", "light", "moderate", "heavy"]:
        sub = df_player[df_player["suspicion_level"] == level]
        if sub.empty:
            continue
        sizes = [dot_size(s) for s in sub["total_shots"]]
        ax.scatter(
            sub["overall_accuracy"], sub["max_accuracy"],
            c=suspicion_color(level), s=sizes,
            alpha=0.75, edgecolors="#ffffff33", linewidths=0.4,
            label=NERF_COLORS[level][1],
        )

    ax.set_xlabel("Átlag pontosság (%) — összes fegyver összesítve")
    ax.set_ylabel("Maximális pontosság (%) — legjobb fegyver")
    ax.grid(True)
    add_legend(ax, ["clean", "light", "moderate", "heavy"])
    add_annotations(ax, df_player, "overall_accuracy", "max_accuracy")

    # Diagonal: both metrics equal
    lim = max(ax.get_xlim()[1], ax.get_ylim()[1])
    ax.plot([0, lim], [0, lim], color="#ffffff22", linestyle="--", linewidth=1)

    ax.text(0.99, 0.02, "Pontméret = lövések száma (több adat = nagyobb pont)",
            transform=ax.transAxes, ha="right", fontsize=7, color="#aaaacc")

    return save_fig(fig, out_dir, "1_accuracy_islands")


def chart_accuracy_vs_ping(df_player: pd.DataFrame, out_dir: Path) -> Path:
    """Accuracy vs ping anomaly count — lagswitch cheaters stand out."""
    fig, ax = plt.subplots(figsize=(10, 7))
    fig.suptitle("Pontosság vs Ping anomáliák (lagswitch gyanú)", color="#e0e0f0", fontsize=13)

    for level in ["clean", "light", "moderate", "heavy"]:
        sub = df_player[df_player["suspicion_level"] == level]
        if sub.empty:
            continue
        sizes = [dot_size(s) for s in sub["total_shots"]]
        ax.scatter(
            sub["overall_accuracy"], sub["ping_anomalies"],
            c=suspicion_color(level), s=sizes,
            alpha=0.75, edgecolors="#ffffff33", linewidths=0.4,
        )

    ax.set_xlabel("Átlag pontosság (%) — összes fegyver")
    ax.set_ylabel("Ping anomáliák száma")
    ax.grid(True)
    add_legend(ax, ["clean", "light", "moderate", "heavy"])

    # Highlight zone: high accuracy + high anomalies = lagswitch candidate
    x_thr = df_player["overall_accuracy"].quantile(0.75) if len(df_player) > 4 else 70
    y_thr = df_player["ping_anomalies"].quantile(0.75) if len(df_player) > 4 else 5
    ax.axvline(x=x_thr, color="#ff990055", linestyle="--", linewidth=1)
    ax.axhline(y=y_thr, color="#ff990055", linestyle="--", linewidth=1)
    ax.text(x_thr + 0.5, y_thr + 0.5, "Lagswitch veszélyzóna",
            color="#ff9900", fontsize=8, alpha=0.8)

    return save_fig(fig, out_dir, "2_accuracy_vs_ping_anomalies")


def chart_kdr_vs_accuracy(df_player: pd.DataFrame, out_dir: Path) -> Path:
    """KDR vs accuracy — cheaters cluster top-right."""
    fig, ax = plt.subplots(figsize=(10, 7))
    fig.suptitle("KDR vs Pontosság szigetek", color="#e0e0f0", fontsize=13)

    for level in ["clean", "light", "moderate", "heavy"]:
        sub = df_player[df_player["suspicion_level"] == level]
        if sub.empty:
            continue
        sizes = [dot_size(s) for s in sub["total_shots"]]
        ax.scatter(
            sub["overall_accuracy"], sub["kdr"],
            c=suspicion_color(level), s=sizes,
            alpha=0.75, edgecolors="#ffffff33", linewidths=0.4,
        )

    ax.set_xlabel("Átlag pontosság (%)")
    ax.set_ylabel("KDR (ölés / halál arány)")
    ax.grid(True)
    add_legend(ax, ["clean", "light", "moderate", "heavy"])
    add_annotations(ax, df_player, "overall_accuracy", "kdr")

    return save_fig(fig, out_dir, "3_kdr_vs_accuracy")


def chart_lagswitch_incidents(df_player: pd.DataFrame, out_dir: Path) -> Path:
    """Lagswitch incidents vs ping stddev — dedicated lagswitch analysis."""
    fig, ax = plt.subplots(figsize=(10, 7))
    fig.suptitle("Lagswitch incidensek vs Ping szórás", color="#e0e0f0", fontsize=13)

    for level in ["clean", "light", "moderate", "heavy"]:
        sub = df_player[df_player["suspicion_level"] == level]
        if sub.empty:
            continue
        sizes = [dot_size(s) for s in sub["total_shots"]]
        ax.scatter(
            sub["ping_stddev"], sub["ls_incidents"],
            c=suspicion_color(level), s=sizes,
            alpha=0.75, edgecolors="#ffffff33", linewidths=0.4,
        )

    ax.set_xlabel("Ping szórás (ms)")
    ax.set_ylabel("Lagswitch incidensek száma")
    ax.grid(True)
    add_legend(ax, ["clean", "light", "moderate", "heavy"])

    return save_fig(fig, out_dir, "4_lagswitch_incidents")


def chart_nerf_distribution(df_player: pd.DataFrame, out_dir: Path) -> Path:
    """Histogram of global nerf % — shows how many players are penalized and how much."""
    fig, ax = plt.subplots(figsize=(10, 6))
    fig.suptitle("Sebzéscsökkentés eloszlása a szerveren", color="#e0e0f0", fontsize=13)

    bins = list(range(0, 105, 5))
    colors_per_bin = []
    for b in bins[:-1]:
        if b == 0:
            colors_per_bin.append(NERF_COLORS["clean"][0])
        elif b <= 30:
            colors_per_bin.append(NERF_COLORS["light"][0])
        elif b <= 60:
            colors_per_bin.append(NERF_COLORS["moderate"][0])
        else:
            colors_per_bin.append(NERF_COLORS["heavy"][0])

    n, _, patches = ax.hist(df_player["nerf_pct"], bins=bins, edgecolor="#ffffff44")
    for patch, color in zip(patches, colors_per_bin):
        patch.set_facecolor(color)
        patch.set_alpha(0.8)

    ax.set_xlabel("Aktív sebzéscsökkentés (%)")
    ax.set_ylabel("Játékosok száma")
    ax.grid(True, axis="y")

    clean_count = (df_player["nerf_pct"] == 0).sum()
    penalized_count = (df_player["nerf_pct"] > 0).sum()
    ax.text(0.98, 0.95,
            f"Normál: {clean_count}  |  Büntetett: {penalized_count}",
            transform=ax.transAxes, ha="right", va="top",
            color="#e0e0f0", fontsize=10,
            bbox=dict(boxstyle="round", facecolor="#3a3a5a", alpha=0.7))

    return save_fig(fig, out_dir, "5_nerf_distribution")


def chart_summary_heatmap(df_player: pd.DataFrame, out_dir: Path) -> Path:
    """Correlation heatmap of key metrics."""
    cols = ["overall_accuracy", "kdr", "ping_avg", "ping_anomalies",
            "ls_incidents", "nerf_pct", "total_shots"]
    labels = ["Átlag\npontosság", "KDR", "Átlag\nping", "Ping\nanoméliák",
              "Lagswitch\nincidensek", "Nerf %", "Lövések\nszáma"]

    available = [c for c in cols if c in df_player.columns]
    avail_labels = [labels[cols.index(c)] for c in available]

    corr = df_player[available].corr()

    fig, ax = plt.subplots(figsize=(9, 8))
    fig.suptitle("Metrikák közötti korreláció", color="#e0e0f0", fontsize=13)

    sns.heatmap(corr, annot=True, fmt=".2f", ax=ax,
                xticklabels=avail_labels, yticklabels=avail_labels,
                cmap="RdYlGn", center=0, vmin=-1, vmax=1,
                linewidths=0.5, linecolor="#1e1e2e",
                cbar_kws={"shrink": 0.8})
    ax.set_facecolor("#2a2a3e")

    return save_fig(fig, out_dir, "6_correlation_heatmap")


# ---------------------------------------------------------------------------
# Excel workbook generation
# ---------------------------------------------------------------------------

EXCEL_DARK_BG   = "1E1E2E"
EXCEL_HEADER_BG = "3A3A5E"
EXCEL_CLEAN     = "2E7D32"
EXCEL_LIGHT     = "F9A825"
EXCEL_MODERATE  = "E65100"
EXCEL_HEAVY     = "B71C1C"

LEVEL_TO_FILL = {
    "clean":    PatternFill("solid", fgColor=EXCEL_CLEAN),
    "light":    PatternFill("solid", fgColor=EXCEL_LIGHT),
    "moderate": PatternFill("solid", fgColor=EXCEL_MODERATE),
    "heavy":    PatternFill("solid", fgColor=EXCEL_HEAVY),
}

def _header_style(ws, row: int, cols: list):
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=EXCEL_HEADER_BG)
    for i, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=i, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")


def build_excel(df_raw: pd.DataFrame, df_player: pd.DataFrame,
                chart_paths: dict, out_path: Path):
    wb = Workbook()

    # ---- Sheet 1: Summary ----
    ws_sum = wb.active
    ws_sum.title = "Összefoglalás"

    total = len(df_player)
    clean = (df_player["suspicion_level"] == "clean").sum()
    light = (df_player["suspicion_level"] == "light").sum()
    moderate = (df_player["suspicion_level"] == "moderate").sum()
    heavy = (df_player["suspicion_level"] == "heavy").sum()
    overridden = df_player["has_override"].sum()

    title_font = Font(bold=True, size=14, color="FFFFFF")
    ws_sum["A1"] = "MogyAntiCheat — Szerver Statisztika Összefoglalás"
    ws_sum["A1"].font = title_font

    rows = [
        ("Összes nyomon követett játékos", total),
        ("Normál játékosok (nincs nerf)", clean),
        ("Enyhe nerf (1–30%)", light),
        ("Mérsékelt nerf (30–60%)", moderate),
        ("Súlyos nerf (>60%)", heavy),
        ("Kézi felülbírálat alatt", int(overridden)),
        ("Átlag pontosság (%)", round(df_player["overall_accuracy"].mean(), 1)),
        ("Átlag KDR", round(df_player["kdr"].mean(), 2)),
        ("Átlag ping (ms)", round(df_player["ping_avg"].mean(), 1)),
    ]

    for r, (label, value) in enumerate(rows, start=3):
        ws_sum.cell(row=r, column=1, value=label).font = Font(color="E0E0F0")
        val_cell = ws_sum.cell(row=r, column=2, value=value)
        val_cell.font = Font(bold=True, color="E0E0F0")

    ws_sum.column_dimensions["A"].width = 36
    ws_sum.column_dimensions["B"].width = 16

    # ---- Sheet 2: Player details ----
    ws_pl = wb.create_sheet("Játékos részletek")
    pl_cols = [
        "player_id", "overall_accuracy", "max_accuracy", "total_shots",
        "kills", "deaths", "assists", "kdr",
        "ping_avg", "ping_stddev", "ping_anomalies",
        "ls_incidents", "nerf_pct", "suspicion_level",
    ]
    col_labels = [
        "Player ID", "Átlag acc%", "Max acc%", "Lövések",
        "Ölések", "Halálok", "Aszisztek", "KDR",
        "Ping átl.", "Ping σ", "Ping anom.",
        "LS incidens", "Nerf %", "Szint",
    ]
    _header_style(ws_pl, 1, col_labels)

    for r, row in enumerate(df_player[pl_cols].itertuples(index=False), start=2):
        for c, val in enumerate(row, start=1):
            cell = ws_pl.cell(row=r, column=c, value=val)
            cell.font = Font(color="E0E0F0")
        # Color-code suspicion level column
        level = df_player.iloc[r - 2]["suspicion_level"]
        ws_pl.cell(row=r, column=14).fill = LEVEL_TO_FILL.get(level, LEVEL_TO_FILL["clean"])

    for i, _ in enumerate(col_labels, start=1):
        ws_pl.column_dimensions[ws_pl.cell(row=1, column=i).column_letter].width = 14

    # ---- Sheet 3: Per-weapon ----
    ws_wpn = wb.create_sheet("Fegyver részletek")
    wpn_cols = ["player_id", "weapon", "accuracy_pct", "shots", "hits", "nerf_pct"]
    wpn_labels = ["Player ID", "Fegyver", "Pontosság %", "Lövések", "Találatok", "Nerf %"]
    _header_style(ws_wpn, 1, wpn_labels)
    for r, row in enumerate(df_raw[["player_id", "weapon", "accuracy_pct", "shots", "hits", "nerf_pct"]].itertuples(index=False), start=2):
        for c, val in enumerate(row, start=1):
            ws_wpn.cell(row=r, column=c, value=round(val, 4) if isinstance(val, float) else val).font = Font(color="E0E0F0")

    # ---- Chart sheets ----
    chart_sheet_map = {
        "1_accuracy_islands": "Pontosság szigetek",
        "2_accuracy_vs_ping_anomalies": "Pontosság vs Ping",
        "3_kdr_vs_accuracy": "KDR vs Pontosság",
        "4_lagswitch_incidents": "Lagswitch",
        "5_nerf_distribution": "Nerf eloszlás",
        "6_correlation_heatmap": "Korreláció",
    }
    for key, sheet_name in chart_sheet_map.items():
        if key not in chart_paths:
            continue
        ws_img = wb.create_sheet(sheet_name)
        img = XLImage(str(chart_paths[key]))
        img.anchor = "A1"
        ws_img.add_image(img)

    wb.save(out_path)
    print(f"  Excel: {out_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="MogyAntiCheat adatelemzés és vizualizáció")
    parser.add_argument("csv", help="A /ac-export csv által generált CSV fájl elérési útja")
    parser.add_argument("--output", "-o", default=None,
                        help="Kimeneti fájl neve (kiterjesztés nélkül). Alapértelmezett: csv neve")
    args = parser.parse_args()

    csv_path = Path(args.csv)
    if not csv_path.exists():
        print(f"Nem találom: {csv_path}")
        sys.exit(1)

    out_name = args.output or csv_path.stem
    out_dir  = csv_path.parent / (out_name + "_charts")
    out_dir.mkdir(exist_ok=True)

    print(f"Beolvasás: {csv_path}")
    df_raw = load_csv(csv_path)
    df_player = player_summary(df_raw)

    total = len(df_player)
    heavy = (df_player["suspicion_level"] == "heavy").sum()
    print(f"  {total} játékos | {heavy} súlyosan gyanús")

    # --- Generate PNG charts ---
    print("Diagramok generálása...")
    chart_paths = {}
    charts = [
        ("1_accuracy_islands",        chart_accuracy_vs_weighted),
        ("2_accuracy_vs_ping_anomalies", chart_accuracy_vs_ping),
        ("3_kdr_vs_accuracy",         chart_kdr_vs_accuracy),
        ("4_lagswitch_incidents",     chart_lagswitch_incidents),
        ("5_nerf_distribution",       chart_nerf_distribution),
        ("6_correlation_heatmap",     chart_summary_heatmap),
    ]
    for key, fn in charts:
        try:
            path = fn(df_player, out_dir)
            chart_paths[key] = path
            print(f"  {path.name}")
        except Exception as e:
            print(f"  [SKIP] {key}: {e}")

    # --- Generate Excel ---
    print("Excel fájl generálása...")
    excel_path = csv_path.parent / (out_name + "_riport.xlsx")
    try:
        build_excel(df_raw, df_player, chart_paths, excel_path)
    except Exception as e:
        print(f"  [HIBA] Excel: {e}")

    print(f"\nKész! Kimenetek: {csv_path.parent / out_name}_charts/")
    print(f"         Excel:  {excel_path}")


if __name__ == "__main__":
    main()
